"""
Excel_automate.py
-----------------
Reads Zerodha order book CSV and updates the portfolio journal (Sample_pf.xlsx).

ORDER BOOK CSV columns used:
  symbol, trade_date, quality (= qty filled), price, trade_type

PORTFOLIO EXCEL layout (Sheet1):
  Row 7  → header row
  Row 8+ → data blocks, each stock occupies exactly BLOCK_SIZE (6) rows:
    Offset 0 (main row)  : Type(A), Symbol(B), Sector(C), Position(D),
                           Tranche=1(E), Date(F), Rec.Amt(G), Rec.Qty(H),
                           Act.Qty(I), Act.Price(J), Curr.Amt(K), Cum.Qty(L), Avg(M)
    Offset 1             : 'Max Allocation:'(A), 'CMP:'(C), Tranche=2(E), Date(F), Act.Qty(I), Act.Price(J)
    Offset 2             : 'Curr Allocation:'(A), 'Curr Profit:'(C), Tranche=3(E), ...
    Offset 3             : 'Curr Qty:'(A), 'Profit%'(C), Tranche=4(E), ...
    Offset 4             : 'Can Add:'(A), 'Booked Profit:'(C), Tranche=5(E), ...
    Offset 5 (meta row)  : Cap type(A), Exchange(B), RR(D), 'Exit on:'(E)

Columns to be filled from order book:
  Type(A), Symbol(B), Sector(C) → manual / blank for new entries
  Position(D)  → default 'Open'
  Date(F)      → trade_date
  Act.Qty(I)   → quality (quantity)
  Act.Price(J) → price

Usage:
  python Excel_automate.py                          # uses defaults below
  python Excel_automate.py --csv my_orders.csv --xl My_Portfolio.xlsx
"""

import argparse
import sys
from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

# ─── Configuration ────────────────────────────────────────────────────────────
DEFAULT_CSV  = "order_book.csv"
DEFAULT_XLSX = "Sample_pf.xlsx"
SHEET_NAME   = "Sheet1"

HEADER_ROW   = 7   # 1-indexed row containing column headers
DATA_START   = 8   # 1-indexed row where first stock block begins
BLOCK_SIZE   = 6   # rows per stock entry
MAX_TRANCHES = 5   # tranches per stock (rows offset 0..4)

# Column indices (1-indexed, matching openpyxl)
COL = {
    "type":     1,   # A
    "symbol":   2,   # B
    "sector":   3,   # C
    "position": 4,   # D
    "tranche":  5,   # E
    "date":     6,   # F
    "rec_amt":  7,   # G
    "rec_qty":  8,   # H
    "act_qty":  9,   # I
    "act_price":10,  # J
    "curr_amt": 11,  # K
    "cum_qty":  12,  # L
    "avg":      13,  # M
}

# Static labels in each row offset within a block (A / C columns)
YELLOW = "FFFFFF00"   # openpyxl RGB for yellow fill

def is_yellow(cell) -> bool:
    """Return True if the cell has a yellow background fill."""
    try:
        return (cell.fill and
                cell.fill.fgColor and
                cell.fill.fgColor.type == "rgb" and
                cell.fill.fgColor.rgb == YELLOW)
    except Exception:
        return False


def write_yellow(cell, value, number_format: str | None = None):
    """Write value only if the cell is yellow; silently skip otherwise."""
    if is_yellow(cell):
        cell.value = value
        if number_format and value is not None:
            cell.number_format = number_format


BLOCK_LABELS = [
    ("Max Allocation:", "CMP:"),          # offset 1
    ("Curr Allocation:", "Curr Profit:"), # offset 2
    ("Curr Qty:", "Profit%"),             # offset 3
    ("Can Add:", "Booked Profit:"),       # offset 4
]

# ─── Helpers ──────────────────────────────────────────────────────────────────

def parse_date(val) -> datetime | None:
    """Parse various date string formats into a datetime object."""
    if pd.isna(val) or val in (None, ""):
        return None
    if isinstance(val, datetime):
        return val
    val = str(val).strip()
    for fmt in ("%m/%d/%Y", "%d-%m-%Y", "%Y-%m-%d", "%m-%d-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    # Try pandas flexible parser as last resort
    try:
        return pd.to_datetime(val).to_pydatetime()
    except Exception:
        return None


def load_orderbook(csv_path: str) -> pd.DataFrame:
    """Load and clean the Zerodha order book CSV."""
    df = pd.read_csv(csv_path)

    # Normalise column names
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

    # Drop completely empty rows
    df.dropna(how="all", inplace=True)

    # Normalise trade_type
    df["trade_type"] = df["trade_type"].str.strip().str.lower()

    # Clean up common columns
    df["symbol"]  = df["symbol"].str.strip().str.upper()
    df["quality"] = pd.to_numeric(df["quality"], errors="coerce").fillna(0).astype(int)
    df["price"]   = pd.to_numeric(df["price"],   errors="coerce")
    df["trade_date_parsed"] = df["trade_date"].apply(parse_date)

    buys  = df[df["trade_type"] == "buy"].copy()
    sells = df[df["trade_type"] == "sell"].copy()

    buys.sort_values(["symbol", "trade_date_parsed"],  inplace=True, ignore_index=True)
    sells.sort_values(["symbol", "trade_date_parsed"], inplace=True, ignore_index=True)

    print(f"✅  Loaded {len(buys)} BUY and {len(sells)} SELL order(s) from '{csv_path}'")
    return buys, sells


def get_stock_blocks(ws) -> tuple[dict, dict, list]:
    """
    Scan the worksheet and return:
      blocks       – dict keyed by UPPER symbol → the LATEST (active) block info
      all_blocks   – dict keyed by UPPER symbol → list of ALL block infos (including overflows)
      empty_slots  – list of start_row values for pre-allocated empty blocks
    Each block_info dict contains:
      main_row, meta_row, tranches, filled_tranches, tranche_keys
    """
    blocks      = {}
    all_blocks  = {}
    empty_slots = []
    row         = DATA_START
    max_row     = ws.max_row

    while row <= max_row:
        symbol_cell  = ws.cell(row=row, column=COL["symbol"]).value
        tranche_cell = ws.cell(row=row, column=COL["tranche"]).value

        if tranche_cell == 1:
            date_cell = ws.cell(row=row, column=COL["date"])
            qty_cell  = ws.cell(row=row, column=COL["act_qty"])

            main_is_empty = (
                is_yellow(date_cell) and date_cell.value in (None, "") and
                is_yellow(qty_cell)  and qty_cell.value  in (None, 0, "")
            )

            if not main_is_empty and symbol_cell not in (None, ""):
                symbol     = str(symbol_cell).strip().upper()
                block_info = {
                    "main_row":        row,
                    "meta_row":        row + 5,
                    "tranches":        {},
                    "filled_tranches": [],
                    "tranche_keys":    set(),
                }
                for offset in range(MAX_TRANCHES):
                    t_row = row + offset
                    t_num = ws.cell(row=t_row, column=COL["tranche"]).value
                    if isinstance(t_num, int) and 1 <= t_num <= MAX_TRANCHES:
                        block_info["tranches"][t_num] = t_row
                        aq = ws.cell(row=t_row, column=COL["act_qty"]).value
                        ap = ws.cell(row=t_row, column=COL["act_price"]).value
                        fd = ws.cell(row=t_row, column=COL["date"]).value
                        if aq not in (None, 0, ""):
                            block_info["filled_tranches"].append(t_num)
                            date_str = fd.strftime("%Y-%m-%d") if isinstance(fd, datetime) else str(fd)
                            block_info["tranche_keys"].add((date_str, int(aq), round(float(ap), 2)))
                # latest active block
                blocks[symbol] = block_info
                # append to full list
                if symbol not in all_blocks:
                    all_blocks[symbol] = []
                all_blocks[symbol].append(block_info)

            else:
                empty_slots.append(row)

            row += BLOCK_SIZE
        else:
            row += 1

    return blocks, all_blocks, empty_slots


def next_empty_tranche(block_info: dict) -> int | None:
    """Return the next tranche number that has no data, or None if all full."""
    filled = set(block_info["filled_tranches"])
    for t in range(1, MAX_TRANCHES + 1):
        if t not in filled:
            return t
    return None


def write_tranche_data(ws, row: int, tranche_num: int,
                       date: datetime | None, act_qty: int, act_price: float):
    """Write date, act_qty, act_price only into yellow cells of the given row."""
    # Tranche number cell is not yellow — write unconditionally (it's a label)
    ws.cell(row=row, column=COL["tranche"]).value = tranche_num
    write_yellow(ws.cell(row=row, column=COL["date"]),      date,      "MM-DD-YYYY")
    write_yellow(ws.cell(row=row, column=COL["act_qty"]),   act_qty)
    write_yellow(ws.cell(row=row, column=COL["act_price"]), act_price)


def append_new_stock_block(ws, symbol: str, date: datetime | None,
                           act_qty: int, act_price: float):
    """
    Append a new 6-row stock block at the bottom, copying yellow cell formatting
    from the very first filled block (row DATA_START) as the reference template.
    Only writes values into yellow cells.
    """
    import copy

    ref_start = DATA_START   # copy formatting from the first block

    # Find first empty row after existing data + 1 blank separator
    start_row = ws.max_row + 1
    while any(ws.cell(row=start_row, column=c).value is not None
              for c in range(1, 14)):
        start_row += 1
    start_row += 1   # blank separator row

    # Copy the 6-row block structure + yellow formatting from reference
    for offset in range(BLOCK_SIZE):
        for col in range(1, 14):
            src  = ws.cell(row=ref_start + offset, column=col)
            dest = ws.cell(row=start_row + offset, column=col)
            # Copy fill (background color)
            if src.fill and src.fill.fgColor:
                dest.fill = copy.copy(src.fill)
            # Copy non-data cell values (labels, tranche numbers) but clear data
            if col not in (COL["symbol"], COL["date"], COL["act_qty"], COL["act_price"]):
                if col != COL["type"]:   # don't copy Type — keep "Satellite"
                    dest.value = src.value

    # Now write the actual stock data into yellow cells only
    write_yellow(ws.cell(row=start_row, column=COL["type"]),     "Satellite")
    write_yellow(ws.cell(row=start_row, column=COL["symbol"]),   symbol)
    write_yellow(ws.cell(row=start_row, column=COL["position"]), "Open")
    write_tranche_data(ws, start_row, 1, date, act_qty, act_price)

    # Clear any data values copied from reference into tranche rows 2-5
    for offset in range(1, MAX_TRANCHES):
        for col in (COL["date"], COL["act_qty"], COL["act_price"]):
            ws.cell(row=start_row + offset, column=col).value = None

    print(f"   ➕  Appended new block for '{symbol}' at row {start_row}")
    return start_row


# ─── Main ─────────────────────────────────────────────────────────────────────

def fill_empty_slot(ws, start_row: int, symbol: str,
                    date: datetime | None, act_qty: int, act_price: float):
    """
    Fill a pre-allocated block at start_row.
    Writes ONLY into yellow cells:
      Main row  → Type(A), Symbol(B), Position(D), Date(F), Act.Qty(I), Act.Price(J)
      Tranche rows → Date(F), Act.Qty(I), Act.Price(J)  (only tranche-1 for new entry)
    Sector(C) is already in the template — user fills it manually.
    """
    write_yellow(ws.cell(row=start_row, column=COL["type"]),     "Satellite")
    write_yellow(ws.cell(row=start_row, column=COL["symbol"]),   symbol)
    write_yellow(ws.cell(row=start_row, column=COL["position"]), "Open")
    write_tranche_data(ws, start_row, 1, date, act_qty, act_price)
    print(f"   📥  Filled pre-allocated slot at row {start_row} for '{symbol}'")


def fix_meta_rows(ws):
    """
    Ensure every block's meta row (offset 5 from the main row) has the correct
    static values: 'Small Cap' in A and 'RR:' in D.  Fixes blocks created by
    older versions of the script that left those cells blank.
    """
    fixed = 0
    row = DATA_START
    while row <= ws.max_row:
        tranche_val = ws.cell(row=row, column=COL["tranche"]).value
        symbol_val  = ws.cell(row=row, column=COL["symbol"]).value

        if tranche_val == 1 and symbol_val not in (None, ""):
            meta_row = row + 5
            a_val = ws.cell(row=meta_row, column=COL["type"]).value
            d_val = ws.cell(row=meta_row, column=COL["position"]).value

            if a_val in (None, ""):
                ws.cell(row=meta_row, column=COL["type"]).value = "Small Cap"
                fixed += 1
            if d_val in (None, ""):
                ws.cell(row=meta_row, column=COL["position"]).value = "RR:"

            row += BLOCK_SIZE
        else:
            row += 1

    if fixed:
        print(f"   🔧  Fixed meta rows for {fixed} existing block(s).")


def main(csv_path: str, xlsx_path: str):
    print(f"\n{'─'*60}")
    print(f"  Portfolio Updater — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'─'*60}")

    # 1. Load order book (returns buys and sells separately)
    buys, sells = load_orderbook(csv_path)
    if buys.empty and sells.empty:
        print("Nothing to update. Exiting.")
        return

    # 2. Load workbook
    wb = openpyxl.load_workbook(xlsx_path)
    if SHEET_NAME not in wb.sheetnames:
        print(f"❌  Sheet '{SHEET_NAME}' not found in '{xlsx_path}'. "
              f"Available: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[SHEET_NAME]
    print(f"✅  Opened workbook '{xlsx_path}' → sheet '{SHEET_NAME}'")

    # 3. Fix any incomplete meta rows from older runs
    fix_meta_rows(ws)

    # 4. Scan existing stock blocks + available empty slots
    blocks, all_blocks, empty_slots = get_stock_blocks(ws)
    print(f"📋  Found {len(blocks)} existing stock(s): {list(blocks.keys()) or 'none'}")
    print(f"🗂️   Found {len(empty_slots)} pre-allocated empty slot(s) available")

    added      = 0
    skipped    = 0
    new_stocks = 0

    # ── 5a. Process BUY orders ────────────────────────────────────────────────
    if not buys.empty:
        print(f"\n{'─'*40}")
        print(f"  Processing {len(buys)} BUY order(s)")
        print(f"{'─'*40}")

    for _, order in buys.iterrows():
        symbol    = order["symbol"]
        act_qty   = int(order["quality"])
        act_price = round(float(order["price"]), 2)
        trade_dt  = order["trade_date_parsed"]

        print(f"\n  BUY: {symbol}  qty={act_qty}  price={act_price}  date={trade_dt}")

        date_str  = trade_dt.strftime("%Y-%m-%d") if isinstance(trade_dt, datetime) else str(trade_dt)
        order_key = (date_str, act_qty, act_price)

        if symbol in blocks:
            block = blocks[symbol]

            if order_key in block["tranche_keys"]:
                print(f"   ⏭  Already in journal. Skipping.")
                skipped += 1
                continue

            t_num = next_empty_tranche(block)

            if t_num is None:
                # All 5 tranches full → overflow into a new block
                print(f"   🔄  All {MAX_TRANCHES} tranches full for '{symbol}'. Opening overflow block...")
                if empty_slots:
                    start = empty_slots.pop(0)
                    fill_empty_slot(ws, start, symbol, trade_dt, act_qty, act_price)
                else:
                    start = append_new_stock_block(ws, symbol, trade_dt, act_qty, act_price)

                overflow_block = {
                    "main_row":        start,
                    "meta_row":        start + 5,
                    "tranches":        {i: start + i - 1 for i in range(1, MAX_TRANCHES + 1)},
                    "filled_tranches": [1],
                    "tranche_keys":    {order_key},
                }
                blocks[symbol] = overflow_block
                if symbol not in all_blocks:
                    all_blocks[symbol] = []
                all_blocks[symbol].append(overflow_block)
                added += 1
                continue

            t_row = block["tranches"].get(t_num)
            if t_row is None:
                print(f"   ⚠  Could not locate tranche row {t_num}. Skipping.")
                skipped += 1
                continue

            write_tranche_data(ws, t_row, t_num, trade_dt, act_qty, act_price)
            block["filled_tranches"].append(t_num)
            block["tranche_keys"].add(order_key)
            print(f"   ✏️  Updated tranche {t_num} for '{symbol}' at row {t_row}")
            added += 1

        else:
            if empty_slots:
                slot_row = empty_slots.pop(0)
                fill_empty_slot(ws, slot_row, symbol, trade_dt, act_qty, act_price)
                start = slot_row
            else:
                start = append_new_stock_block(ws, symbol, trade_dt, act_qty, act_price)

            new_block = {
                "main_row":        start,
                "meta_row":        start + 5,
                "tranches":        {i: start + i - 1 for i in range(1, MAX_TRANCHES + 1)},
                "filled_tranches": [1],
                "tranche_keys":    {order_key},
            }
            blocks[symbol]    = new_block
            all_blocks[symbol] = [new_block]
            added += 1
            new_stocks += 1

    # ── 5b. Process SELL orders ───────────────────────────────────────────────
    if not sells.empty:
        print(f"\n{'─'*40}")
        print(f"  Processing {len(sells)} SELL order(s)")
        print(f"{'─'*40}")

    for _, order in sells.iterrows():
        symbol    = order["symbol"]
        sell_qty  = int(order["quality"])
        sell_price = round(float(order["price"]), 2)
        sell_dt   = order["trade_date_parsed"]

        print(f"\n  SELL: {symbol}  qty={sell_qty}  price={sell_price}  date={sell_dt}")

        if symbol not in all_blocks:
            print(f"   ⚠  '{symbol}' not found in journal. Skipping sell.")
            skipped += 1
            continue

        # Write exit date + price into the meta row of ALL blocks for this symbol.
        # Position → "Closed" on ALL blocks (including overflows).
        # Overflow blocks (index 1+) get only exit date + price in meta row — no other changes.
        sym_blocks = all_blocks[symbol]
        for idx, blk in enumerate(sym_blocks):
            meta  = blk["meta_row"]
            main  = blk["main_row"]

            # Exit date → F cell of meta row
            meta_f = ws.cell(row=meta, column=COL["date"])
            meta_f.value = sell_dt
            meta_f.number_format = "MM-DD-YYYY"

            # Exit price → J cell of meta row
            meta_j = ws.cell(row=meta, column=COL["act_price"])
            meta_j.value = sell_price

            # Position → "Closed" on all blocks
            write_yellow(ws.cell(row=main, column=COL["position"]), "Closed")

        print(f"   🔴  Marked '{symbol}' as CLOSED | exit date={sell_dt} | exit price={sell_price}")
        added += 1

    # 6. Save workbook
    wb.save(xlsx_path)
    print(f"\n{'─'*60}")
    print(f"  Done! {added} write(s) | {new_stocks} new stock(s) | {skipped} skipped")
    print(f"  Saved → '{xlsx_path}'")
    print(f"{'─'*60}\n")


# ─── Stream-based entry point (used by the Streamlit UI) ─────────────────────

def process_streams(csv_bytes: bytes, xlsx_bytes: bytes) -> tuple[bytes, list[str]]:
    """
    Same logic as main() but works entirely in memory.
    Accepts raw bytes for both files and returns:
      (updated_xlsx_bytes, log_lines)
    """
    import io

    log: list[str] = []

    def log_print(*args):
        msg = " ".join(str(a) for a in args)
        log.append(msg)

    # 1. Load order book from bytes
    csv_io = io.BytesIO(csv_bytes)
    df = pd.read_csv(csv_io)
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    df.dropna(how="all", inplace=True)
    df["trade_type"] = df["trade_type"].str.strip().str.lower()
    df["symbol"]     = df["symbol"].str.strip().str.upper()
    df["quality"]    = pd.to_numeric(df["quality"], errors="coerce").fillna(0).astype(int)
    df["price"]      = pd.to_numeric(df["price"],   errors="coerce")
    df["trade_date_parsed"] = df["trade_date"].apply(parse_date)

    buys  = df[df["trade_type"] == "buy"].copy().sort_values(
        ["symbol", "trade_date_parsed"], ignore_index=True)
    sells = df[df["trade_type"] == "sell"].copy().sort_values(
        ["symbol", "trade_date_parsed"], ignore_index=True)

    log_print(f"✅ Loaded {len(buys)} BUY and {len(sells)} SELL order(s)")

    if buys.empty and sells.empty:
        log_print("⚠ No orders to process.")
        return xlsx_bytes, log

    # 2. Load workbook from bytes
    xlsx_io = io.BytesIO(xlsx_bytes)
    wb = openpyxl.load_workbook(xlsx_io)
    if SHEET_NAME not in wb.sheetnames:
        log_print(f"❌ Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        return xlsx_bytes, log
    ws = wb[SHEET_NAME]

    fix_meta_rows(ws)
    blocks, all_blocks, empty_slots = get_stock_blocks(ws)
    log_print(f"📋 Found {len(blocks)} existing stock(s): {list(blocks.keys()) or 'none'}")
    log_print(f"🗂️ Found {len(empty_slots)} empty slot(s) available")

    added = skipped = new_stocks = 0

    # ── BUY orders ────────────────────────────────────────────────────────────
    for _, order in buys.iterrows():
        symbol    = order["symbol"]
        act_qty   = int(order["quality"])
        act_price = round(float(order["price"]), 2)
        trade_dt  = order["trade_date_parsed"]
        date_str  = trade_dt.strftime("%Y-%m-%d") if isinstance(trade_dt, datetime) else str(trade_dt)
        order_key = (date_str, act_qty, act_price)

        if symbol in blocks:
            block = blocks[symbol]
            if order_key in block["tranche_keys"]:
                log_print(f"  ⏭ BUY {symbol} {date_str} qty={act_qty} — already in journal, skipped")
                skipped += 1
                continue

            t_num = next_empty_tranche(block)
            if t_num is None:
                log_print(f"  🔄 {symbol} all tranches full — overflow block")
                if empty_slots:
                    start = empty_slots.pop(0)
                    fill_empty_slot(ws, start, symbol, trade_dt, act_qty, act_price)
                else:
                    start = append_new_stock_block(ws, symbol, trade_dt, act_qty, act_price)
                overflow_block = {
                    "main_row": start, "meta_row": start + 5,
                    "tranches": {i: start + i - 1 for i in range(1, MAX_TRANCHES + 1)},
                    "filled_tranches": [1], "tranche_keys": {order_key},
                }
                blocks[symbol] = overflow_block
                if symbol not in all_blocks:
                    all_blocks[symbol] = []
                all_blocks[symbol].append(overflow_block)
                added += 1
                continue

            t_row = block["tranches"].get(t_num)
            if t_row is None:
                log_print(f"  ⚠ {symbol} tranche row {t_num} not found, skipped")
                skipped += 1
                continue

            write_tranche_data(ws, t_row, t_num, trade_dt, act_qty, act_price)
            block["filled_tranches"].append(t_num)
            block["tranche_keys"].add(order_key)
            log_print(f"  ✏️ BUY {symbol} tranche {t_num} — qty={act_qty} price={act_price} date={date_str}")
            added += 1

        else:
            if empty_slots:
                slot_row = empty_slots.pop(0)
                fill_empty_slot(ws, slot_row, symbol, trade_dt, act_qty, act_price)
                start = slot_row
            else:
                start = append_new_stock_block(ws, symbol, trade_dt, act_qty, act_price)
            new_block = {
                "main_row": start, "meta_row": start + 5,
                "tranches": {i: start + i - 1 for i in range(1, MAX_TRANCHES + 1)},
                "filled_tranches": [1], "tranche_keys": {order_key},
            }
            blocks[symbol]     = new_block
            all_blocks[symbol] = [new_block]
            log_print(f"  ➕ NEW {symbol} — qty={act_qty} price={act_price} date={date_str}")
            added += 1
            new_stocks += 1

    # ── SELL orders ───────────────────────────────────────────────────────────
    for _, order in sells.iterrows():
        symbol     = order["symbol"]
        sell_price = round(float(order["price"]), 2)
        sell_dt    = order["trade_date_parsed"]

        if symbol not in all_blocks:
            log_print(f"  ⚠ SELL {symbol} — not found in journal, skipped")
            skipped += 1
            continue

        for blk in all_blocks[symbol]:
            meta = blk["meta_row"]
            main = blk["main_row"]
            ws.cell(row=meta, column=COL["date"]).value        = sell_dt
            ws.cell(row=meta, column=COL["date"]).number_format = "MM-DD-YYYY"
            ws.cell(row=meta, column=COL["act_price"]).value   = sell_price
            write_yellow(ws.cell(row=main, column=COL["position"]), "Closed")

        log_print(f"  🔴 SELL {symbol} CLOSED — exit date={sell_dt} exit price={sell_price}")
        added += 1

    # 3. Save to bytes
    out_io = io.BytesIO()
    wb.save(out_io)
    out_io.seek(0)

    log_print(f"\n✅ Done — {added} write(s) | {new_stocks} new | {skipped} skipped")
    return out_io.read(), log


# ─── Entry point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Sync Zerodha order book CSV into portfolio journal XLSX"
    )
    parser.add_argument(
        "--csv", default=DEFAULT_CSV,
        help=f"Path to Zerodha order book CSV (default: {DEFAULT_CSV})"
    )
    parser.add_argument(
        "--xl", default=DEFAULT_XLSX,
        help=f"Path to portfolio journal XLSX (default: {DEFAULT_XLSX})"
    )
    args = parser.parse_args()
    main(args.csv, args.xl)
